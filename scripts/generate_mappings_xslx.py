import os
import re
import itertools
import sys
from collections import defaultdict
from typing import List, Dict, Tuple, Any

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# --- Configuration ---
DEFAULT_YARN_TINY_PATH = r"yarn-1.21.6+build.1-tiny"
DEFAULT_MOJMAP_PATH = r"server.txt"
DEFAULT_OUTPUT_FILENAME = "minecraft-mappings-mojmap-fabric-1.21.6.xlsx"
CLIENT_PACKAGE_PREFIXES = ('net.minecraft.client', 'com.mojang.blaze3d')

# --- Regex & Mappings ---
DESCRIPTOR_RE = re.compile(r'(\[*(?:[BCDFIJSZ]|L[\w/$]+;))')
DESCRIPTOR_TYPE_MAP = {
    'B': 'byte', 'C': 'char', 'D': 'double', 'F': 'float', 'I': 'int',
    'J': 'long', 'S': 'short', 'Z': 'boolean', 'V': 'void'
}
LINE_NUMBER_PREFIX_RE = re.compile(r"^(?:\d+:\d*:|\d+:)")
COLOR_PALETTE = [
    "E0F2F7", "E8F5E9", "FFF8E1", "F3E5F5", "FCE4EC", "FFCCBC",
    "CFD8DC", "D7CCC8", "DCEDC8", "B2EBF2", "FFE0B2", "D1C4E9",
    "C8E6C9", "FFCDD2", "BBDEFB", "FFFDE7", "F0F4C3", "ECCBCC"
]
HEADER_TYPE_COLORS = {
    "Class": "4F81BD", "Field": "C0504D",
    "Method": "9BBB59", "Generic": "808080"
}


class ColorManager:
    """Manages a recycling color palette for Excel rows."""
    def __init__(self, palette: List[str]):
        self._palette_cycler = itertools.cycle(palette)
        self._class_fill_map: Dict[str, PatternFill] = {}

    def get_fill_for_class(self, class_name: str) -> PatternFill:
        if not class_name:
            return None
        if class_name not in self._class_fill_map:
            color_hex = next(self._palette_cycler)
            self._class_fill_map[class_name] = PatternFill(
                start_color=color_hex, end_color=color_hex,
                fill_type="solid"
            )
        return self._class_fill_map[class_name]


def convert_path_to_dot_notation(path: str) -> str:
    """Converts a Java-style path with slashes to dot notation."""
    return path.replace('/', '.')


def parse_descriptor_type(desc: str, class_lookup: Dict[str, str]) -> str:
    """Converts a descriptor part to a readable type, using a lookup."""
    array_suffix = ''
    while desc.startswith('['):
        array_suffix += '[]'
        desc = desc[1:]
    if desc.startswith('L') and desc.endswith(';'):
        path_inside = desc[1:-1]
        deobf_path = class_lookup.get(path_inside, path_inside)
        return convert_path_to_dot_notation(deobf_path) + array_suffix
    return DESCRIPTOR_TYPE_MAP.get(desc, 'unknown') + array_suffix


def parse_yarn_signature(
    descriptor: str, method_name: str, class_lookup: Dict[str, str]
) -> Tuple[str, str]:
    """Parses a Yarn signature using a class lookup for readable types."""
    try:
        match = re.match(r'\((.*)\)(.*)', descriptor)
        if not match:
            return "unknown", "()"
        param_part, return_part = match.groups()
        return_type = parse_descriptor_type(return_part, class_lookup)
        param_types = DESCRIPTOR_RE.findall(param_part)
        param_list = [
            f"{parse_descriptor_type(p_type, class_lookup)} arg{i}"
            for i, p_type in enumerate(param_types)
        ]
        return return_type, f"{method_name}({', '.join(param_list)})"
    except Exception:
        return "error", "()"


def parse_mojmap_signature(full_signature: str) -> Tuple[str, str]:
    """Parses a full Mojang deobfuscated signature."""
    try:
        parts = full_signature.split('(', 1)
        before_params = parts[0].strip()
        params = f"({parts[1]}" if len(parts) > 1 else "()"
        last_space = before_params.rfind(' ')
        if last_space == -1:
            return "", convert_path_to_dot_notation(full_signature)
        return_type = convert_path_to_dot_notation(before_params[:last_space])
        method_name_and_params = convert_path_to_dot_notation(
            f"{before_params[last_space+1:]}{params}"
        )
        return return_type, method_name_and_params
    except Exception:
        return "error", full_signature


def parse_yarn_tiny_file(
    content: str
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """Parses the primary Yarn .tiny mapping file."""
    print("Parsing Yarn .tiny mapping file...")
    classes, fields, methods = [], [], []
    lines = content.splitlines()

    combined_class_lookup = {}
    for line in lines:
        parts = line.strip().split('\t')
        if not parts or parts[0] != 'CLASS':
            continue
        obfuscated, intermediary, yarn = parts[1], parts[2], parts[3]
        combined_class_lookup[intermediary] = yarn
        combined_class_lookup[obfuscated] = yarn

    official_to_class_data = {}
    for line in lines:
        parts = line.strip().split('\t')
        if not parts or parts[0] != 'CLASS':
            continue
        class_info = {
            'Mojmap Obfuscated Class': convert_path_to_dot_notation(parts[1]),
            'Yarn Intermediary Class': convert_path_to_dot_notation(parts[2]),
            'Yarn Deobfuscated Class': convert_path_to_dot_notation(parts[3]),
            'Mojmap Deobfuscated Class': '',
        }
        classes.append(class_info)
        official_to_class_data[class_info['Mojmap Obfuscated Class']] = \
            class_info

    for line in lines:
        parts = line.strip().split('\t')
        if not parts or parts[0] not in ('FIELD', 'METHOD'):
            continue
        keyword, owner_obf, desc, obf, inter, yarn = parts
        owner_data = official_to_class_data.get(
            convert_path_to_dot_notation(owner_obf)
        )
        if not owner_data:
            continue
        base_info = {
            'Mojmap Obfuscated Class': owner_data['Mojmap Obfuscated Class'],
            'Yarn Deobfuscated Class': owner_data['Yarn Deobfuscated Class'],
            'Yarn Intermediary Class': owner_data['Yarn Intermediary Class'],
            'Mojmap Deobfuscated Class': '',
            'Mojmap Obfuscated Field' if keyword == 'FIELD'
            else 'Mojmap Obfuscated Method': obf,
            'Yarn Deobfuscated Field' if keyword == 'FIELD'
            else 'Yarn Deobfuscated Method': yarn,
            'Yarn Intermediary Field' if keyword == 'FIELD'
            else 'Yarn Intermediary Method': inter,
            'Mojmap Deobfuscated Field' if keyword == 'FIELD'
            else 'Mojmap Deobfuscated Method': '',
        }
        if keyword == 'FIELD':
            base_info['Yarn Type'] = parse_descriptor_type(
                desc, combined_class_lookup
            )
            base_info['Mojmap Type'] = ''
            fields.append(base_info)
        elif keyword == 'METHOD':
            ret_type, sig = parse_yarn_signature(
                desc, yarn, combined_class_lookup
            )
            base_info['Yarn Return Type'] = ret_type
            base_info['Yarn Signature'] = sig
            base_info['Mojmap Return Type'] = ''
            base_info['Mojmap Signature'] = ''
            methods.append(base_info)
    print("Yarn parsing complete.")
    return classes, fields, methods


def parse_mojmap_proguard_file(content: str) -> Tuple[Dict, Dict, Dict]:
    """Parses a Mojang ProGuard file into lookup dictionaries."""
    print("Parsing Mojmap ProGuard mapping file...")
    class_map, field_map, method_map = {}, {}, {}
    current_class_obf = ""
    for line in content.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith('#'):
            continue
        processed = LINE_NUMBER_PREFIX_RE.sub('', stripped)
        if processed.endswith(':'):
            parts = processed.rstrip(':').split(' -> ')
            if len(parts) == 2:
                mojmap_name, obf_name = (
                    convert_path_to_dot_notation(parts[0]),
                    convert_path_to_dot_notation(parts[1])
                )
                current_class_obf = obf_name
                class_map[obf_name] = mojmap_name
        elif line.startswith(' '):
            if not current_class_obf:
                continue
            parts = processed.split(' -> ')
            if len(parts) == 2:
                deobf_sig, obf_name = parts[0].strip(), parts[1].strip()
                if '(' in deobf_sig and ')' in deobf_sig:
                    method_map[(current_class_obf, obf_name)] = deobf_sig
                else:
                    field_map[(current_class_obf, obf_name)] = deobf_sig
    print("Mojmap parsing complete.")
    return class_map, field_map, method_map


def merge_mappings(all_data: Tuple, mojmap_maps: Tuple):
    """Merges Mojmap names into the primary data lists."""
    print("Merging Mojmap and Yarn data...")
    classes, fields, methods = all_data
    mojmap_classes, mojmap_fields, mojmap_methods = mojmap_maps
    for item in classes:
        item['Mojmap Deobfuscated Class'] = mojmap_classes.get(
            item['Mojmap Obfuscated Class'], ''
        )
    obf_to_mojmap_owner = {
        c['Mojmap Obfuscated Class']: c['Mojmap Deobfuscated Class']
        for c in classes
    }
    for item in fields:
        field_key = (
            item['Mojmap Obfuscated Class'], item['Mojmap Obfuscated Field']
        )
        full_mojmap_field = mojmap_fields.get(field_key, '')
        item['Mojmap Deobfuscated Class'] = obf_to_mojmap_owner.get(
            item['Mojmap Obfuscated Class'], ''
        )
        if full_mojmap_field:
            field_parts = full_mojmap_field.split()
            item['Mojmap Deobfuscated Field'] = field_parts[-1]
            item['Mojmap Type'] = convert_path_to_dot_notation(
                " ".join(field_parts[:-1])
            )
    for item in methods:
        method_key = (
            item['Mojmap Obfuscated Class'], item['Mojmap Obfuscated Method']
        )
        full_mojmap_sig = mojmap_methods.get(method_key, '')
        item['Mojmap Deobfuscated Method'] = (
            full_mojmap_sig.split('(', 1)[0].split()[-1] if full_mojmap_sig
            else ''
        )
        item['Mojmap Deobfuscated Class'] = obf_to_mojmap_owner.get(
            item['Mojmap Obfuscated Class'], ''
        )
        if full_mojmap_sig:
            ret_type, sig = parse_mojmap_signature(full_mojmap_sig)
            item['Mojmap Return Type'] = ret_type
            item['Mojmap Signature'] = sig
    print("Merging complete.")


def filter_client_packages(all_data: Tuple) -> Tuple:
    """Removes classes and members from client-only packages."""
    print(f"\nFiltering out client packages: {CLIENT_PACKAGE_PREFIXES}")
    classes, fields, methods = all_data
    initial_class_count = len(classes)
    server_classes = [
        c for c in classes
        if not c['Yarn Deobfuscated Class'].startswith(CLIENT_PACKAGE_PREFIXES)
    ]
    kept_class_yarn_names = {
        c['Yarn Deobfuscated Class'] for c in server_classes
    }
    server_fields = [
        f for f in fields
        if f['Yarn Deobfuscated Class'] in kept_class_yarn_names
    ]
    server_methods = [
        m for m in methods
        if m['Yarn Deobfuscated Class'] in kept_class_yarn_names
    ]
    print(
        f"Removed {initial_class_count - len(server_classes)} client classes."
    )
    return server_classes, server_fields, server_methods


def setup_sheet(
    workbook: Workbook, sheet_name: str, headers: List[str]
) -> Worksheet:
    """Sets up a sheet with dynamically colored and formatted headers."""
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_border = Border(bottom=Side(style='thin', color='000000'))
    header_alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True
    )
    sheet.row_dimensions[1].height = 35
    for cell in sheet[1]:
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
        header_text = cell.value
        color = HEADER_TYPE_COLORS["Generic"]
        if "Class" in header_text:
            color = HEADER_TYPE_COLORS["Class"]
        elif "Field" in header_text:
            color = HEADER_TYPE_COLORS["Field"]
        elif "Method" in header_text:
            color = HEADER_TYPE_COLORS["Method"]
        cell.fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
    sheet.freeze_panes = 'A2'
    return sheet


def auto_size_all_columns(sheet: Worksheet):
    """Auto-sizes all columns in a worksheet based on content."""
    column_widths = defaultdict(int)
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value:
                cell_lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in cell_lines)
                column_widths[i] = max(column_widths[i], max_line_length)

    for i, column_width in column_widths.items():
        adjusted_width = min(column_width + 2, 60)
        sheet.column_dimensions[get_column_letter(i + 1)].width = \
            adjusted_width


def write_data_to_sheet(
    sheet: Worksheet, data: List[Dict], color_manager: ColorManager,
    row_fill_key: str
):
    """Writes a list of dictionaries to a sheet with row coloring."""
    headers = [cell.value for cell in sheet[1]]
    print(f"Writing {len(data)} rows to '{sheet.title}' sheet...")

    for row_data in data:
        fill = color_manager.get_fill_for_class(row_data.get(row_fill_key))
        row_to_append = [row_data.get(key, '') for key in headers]
        sheet.append(row_to_append)
        if fill:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill


def create_excel_spreadsheet(
    classes: List[Dict], fields: List[Dict], methods: List[Dict],
    filename: str
):
    """Creates the final Excel file with all data and formatting."""
    print(f"\nCreating Excel spreadsheet: {filename}")
    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        color_manager = ColorManager(COLOR_PALETTE)
        for cls_entry in classes:
            color_manager.get_fill_for_class(
                cls_entry['Yarn Deobfuscated Class']
            )

        class_headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Mojmap Deobfuscated Class", "Yarn Deobfuscated Class"
        ]
        field_headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Mojmap Deobfuscated Class", "Yarn Deobfuscated Class",
            "Mojmap Obfuscated Field", "Yarn Intermediary Field",
            "Mojmap Deobfuscated Field", "Yarn Deobfuscated Field",
            "Mojmap Type", "Yarn Type"
        ]
        method_headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Mojmap Deobfuscated Class", "Yarn Deobfuscated Class",
            "Mojmap Obfuscated Method", "Yarn Intermediary Method",
            "Mojmap Deobfuscated Method", "Yarn Deobfuscated Method",
            "Mojmap Return Type", "Yarn Return Type",
            "Mojmap Signature", "Yarn Signature"
        ]

        classes_sheet = setup_sheet(workbook, "Classes", class_headers)
        write_data_to_sheet(
            classes_sheet, classes, color_manager, "Yarn Deobfuscated Class"
        )

        fields_sheet = setup_sheet(workbook, "Fields", field_headers)
        write_data_to_sheet(
            fields_sheet, fields, color_manager, "Yarn Deobfuscated Class"
        )

        methods_sheet = setup_sheet(workbook, "Methods", method_headers)
        write_data_to_sheet(
            methods_sheet, methods, color_manager, "Yarn Deobfuscated Class"
        )

        print("\nAuto-sizing columns for optimal spacing...")
        for sheet in workbook.worksheets:
            auto_size_all_columns(sheet)
        print("Column sizing complete.")

        workbook.save(filename)
        print(f"\nSuccessfully created '{os.path.abspath(filename)}'")
    except PermissionError:
        print(
            f"\nError: Permission denied. Could not save '{filename}'. "
            "Ensure the file is not open in another program."
        )
    except Exception as e:
        print(f"\nError creating Excel spreadsheet: {e}")


def get_file_content(
    prompt_message: str, default_path: str, is_optional: bool = False
) -> str:
    """Prompts user for a file path and returns its content."""
    prompt = f"{prompt_message} (or press Enter for default):\n> "
    if is_optional:
        prompt = f"{prompt_message} (Optional, press Enter to skip):\n> "
    print(prompt, default_path)
    file_path = input("> ").strip() or default_path
    if not file_path and is_optional:
        return None
    try:
        print(f"Reading from: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        print(f"\nError: File not found at '{file_path}'")
        return None
    except Exception as e:
        print(f"\nError reading file: {e}")
        return None


def get_data_percentage() -> int:
    """Prompts the user for a percentage of data to process."""
    while True:
        p_input = input(
            "\nEnter data percentage to process (1-100), "
            "or -1 for all [default: -1]: "
        ).strip()
        if not p_input:
            return -1
        try:
            percentage = int(p_input)
            if percentage == -1 or 1 <= percentage <= 100:
                return percentage
            else:
                print("Invalid range. Please enter a number between 1-100.")
        except ValueError:
            print("Invalid input. Please enter a number.")


def main():
    """Main script execution flow."""
    script_dir = os.path.dirname(os.path.abspath(__file__))

    yarn_content = get_file_content(
        "Enter path to YARN .tiny mapping file", DEFAULT_YARN_TINY_PATH
    )
    if not yarn_content:
        sys.exit(1)

    all_data = parse_yarn_tiny_file(yarn_content)

    mojmap_content = get_file_content(
        "Enter path to MOJMAP server.txt mapping file", DEFAULT_MOJMAP_PATH,
        is_optional=True
    )
    if mojmap_content:
        mojmap_maps = parse_mojmap_proguard_file(mojmap_content)
        merge_mappings(all_data, mojmap_maps)

    classes, fields, methods = filter_client_packages(all_data)

    print("\nSorting data for improved readability...")
    classes.sort(key=lambda item: item['Yarn Deobfuscated Class'])
    fields.sort(
        key=lambda item: (
            item['Yarn Deobfuscated Class'], item['Yarn Deobfuscated Field']
        )
    )
    methods.sort(
        key=lambda item: (
            item['Yarn Deobfuscated Class'], item['Yarn Deobfuscated Method']
        )
    )
    print("Sorting complete.")

    data_percentage = get_data_percentage()
    if 1 <= data_percentage <= 100:
        print(
            f"\nProcessing top {data_percentage}% of server-side classes..."
        )
        num_classes = int(len(classes) * (data_percentage / 100.0))
        classes = classes[:num_classes]
        kept_names = {c['Yarn Deobfuscated Class'] for c in classes}
        fields = [
            f for f in fields
            if f['Yarn Deobfuscated Class'] in kept_names
        ]
        methods = [
            m for m in methods
            if m['Yarn Deobfuscated Class'] in kept_names
        ]
        print(
            f"Reduced dataset to: {len(classes)} classes, "
            f"{len(fields)} fields, {len(methods)} methods."
        )
    else:
        print("\nProcessing all server-side data.")

    version_part = re.search(r'(\d+\.\d+(\.\d+)?)',
                             os.path.basename(DEFAULT_YARN_TINY_PATH))
    base_name = (
        f"minecraft-mappings-mojmap-fabric-{version_part.group(1)}.xlsx"
        if version_part else DEFAULT_OUTPUT_FILENAME
    )

    print(f"\nThe output file will be saved in: {script_dir}")
    filename = input(
        f"Enter output Excel filename (or press Enter for '{base_name}'): "
    ).strip() or base_name
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    final_path = os.path.join(script_dir, filename)

    create_excel_spreadsheet(classes, fields, methods, final_path)


if __name__ == "__main__":
    main()