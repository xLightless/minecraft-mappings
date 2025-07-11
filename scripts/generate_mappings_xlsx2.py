"""
This script converts the mojang 1.21.6 and fabric (yarn) mappings into an
organized excel spreadsheet. It creates a symmetrical set of 9 sheets:
3 for fully matched classes, fields, and methods; 3 for Yarn entries
without a Mojmap pair; and 3 for Mojmap entries without a Yarn pair.
Color coding is used to associate members back to their respective classes.
"""

import itertools
import os
import re
import sys
from collections import defaultdict
from typing import Dict, List, Optional, Set, Tuple

# This script requires the 'openpyxl' package.
# You can install it with: pip install openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# --- Configuration Constants ---
DEFAULT_YARN_TINY_PATH = (r"C:\Users\xLightless\source\repos\xLightless"
                          r"\WorldManager\net.minecraft.maps\1.21.6"
                          r"\yarn-1.21.6+build.1-tiny")
DEFAULT_MOJMAP_PATH = (r"C:\Users\xLightless\source\repos\xLightless"
                       r"\WorldManager\net.minecraft.maps\1.21.6\server.txt")
DEFAULT_OUTPUT_FILENAME = "mappings-by-category-1.21.6.xlsx"
CLIENT_PACKAGE_PREFIXES = ('net.minecraft.client', 'com.mojang.blaze3d')

# --- Pre-compiled Regex and Static Mappings ---
DESCRIPTOR_RE = re.compile(r'(\[*(?:[BCDFIJSZ]|L[\w/$]+;))')
DESCRIPTOR_TYPE_MAP = {
    'B': 'byte', 'C': 'char', 'D': 'double', 'F': 'float', 'I': 'int',
    'J': 'long', 'S': 'short', 'Z': 'boolean', 'V': 'void'
}
LINE_NUMBER_PREFIX_RE = re.compile(r"^(?:\d+:\d*:|\d+:)")

# --- Excel Formatting Constants ---
COLOR_PALETTE = [
    "E0F2F7", "E8F5E9", "FFF8E1", "F3E5F5", "FCE4EC", "FFCCBC",
    "CFD8DC", "D7CCC8", "DCEDC8", "B2EBF2", "FFE0B2", "D1C4E9",
    "C8E6C9", "FFCDD2", "BBDEFB", "FFFDE7", "F0F4C3", "ECCBCC"
]
HEADER_TYPE_COLORS = {
    "Class": "4F81BD", "Field": "C0504D",
    "Method": "9BBB59", "Generic": "808080"
}


class ColorManager:
    """Manages a recycling color palette for Excel rows."""

    def __init__(self, palette: List[str]):
        """Initializes the color manager with a list of hex color strings."""
        self._palette_cycler = itertools.cycle(palette)
        self._class_fill_map: Dict[str, PatternFill] = {}

    def get_fill_for_class(self, class_name: str) -> Optional[PatternFill]:
        """
        Provides a consistent PatternFill for a given class name.
        A new color is assigned from the palette if the class name is new.
        """
        if not class_name:
            return None
        if class_name not in self._class_fill_map:
            color_hex = next(self._palette_cycler)
            self._class_fill_map[class_name] = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )
        return self._class_fill_map[class_name]


class MappingProcessor:
    """
    Handles loading, parsing, merging, and processing of Minecraft mappings.
    """

    def __init__(self):
        """Initializes storage for mapping data."""
        self.all_yarn_classes: List[Dict] = []
        self.all_yarn_fields: List[Dict] = []
        self.all_yarn_methods: List[Dict] = []

        self.mojmap_classes: Dict[str, str] = {}
        self.mojmap_fields: Dict[Tuple[str, str], str] = {}
        self.mojmap_methods: Dict[Tuple[str, str, str], str] = {}

        self.used_mojmap_class_keys: Set[str] = set()
        self.used_mojmap_field_keys: Set[Tuple[str, str]] = set()
        self.used_mojmap_method_keys: Set[Tuple[str, str, str]] = set()

    # --- Static Helper Methods ---
    @staticmethod
    def _convert_path(path: str) -> str:
        """Converts a Java-style path with slashes to dot notation."""
        return path.replace('/', '.')

    @staticmethod
    def _parse_mojmap_signature(full_sig: str) -> Tuple[str, str]:
        """Parses a Mojang deobfuscated signature into return type and name."""
        try:
            parts = full_sig.split('(', 1)
            before_params = parts[0].strip()
            params = f"({parts[1]}" if len(parts) > 1 else "()"
            last_space = before_params.rfind(' ')
            if last_space == -1:
                return "", MappingProcessor._convert_path(full_sig)

            return_type = MappingProcessor._convert_path(
                before_params[:last_space]
            )
            method_name_and_params = MappingProcessor._convert_path(
                f"{before_params[last_space+1:]}{params}"
            )
            return return_type, method_name_and_params
        except Exception:
            return "error", full_sig

    @staticmethod
    def _parse_descriptor_type(desc: str, lookup: Dict) -> str:
        """Converts a bytecode descriptor part to a readable type name."""
        array_suffix = ''
        while desc.startswith('['):
            array_suffix += '[]'
            desc = desc[1:]

        if desc.startswith('L') and desc.endswith(';'):
            path_inside = desc[1:-1]
            deobf_path = lookup.get(path_inside, path_inside)
            return MappingProcessor._convert_path(deobf_path) + array_suffix

        return DESCRIPTOR_TYPE_MAP.get(desc, 'unknown') + array_suffix

    # --- File Parsing Methods ---
    def parse_yarn_tiny_file(self, content: str):
        """Parses a Yarn .tiny mapping file into instance attributes."""
        print("Parsing Yarn .tiny mapping file...")
        lines = content.splitlines()

        class_lookup = {}
        for line in lines:
            parts = line.strip().split('\t')
            if parts and parts[0] == 'CLASS':
                obf, inter, yarn = parts[1], parts[2], parts[3]
                class_lookup[inter] = yarn
                class_lookup[obf] = yarn

        official_to_class_data = {}
        for line in lines:
            parts = line.strip().split('\t')
            if not parts:
                continue
            if parts[0] == 'CLASS':
                class_info = {
                    'Mojmap Obfuscated Class': self._convert_path(parts[1]),
                    'Yarn Intermediary Class': self._convert_path(parts[2]),
                    'Yarn Deobfuscated Class': self._convert_path(parts[3]),
                    'Mojmap Deobfuscated Class': '',
                }
                self.all_yarn_classes.append(class_info)
                official_to_class_data[class_info['Mojmap Obfuscated Class']] \
                    = class_info
            elif parts[0] in ('FIELD', 'METHOD'):
                self._process_yarn_member(parts, official_to_class_data,
                                          class_lookup)
        print("Yarn parsing complete.")

    def _process_yarn_member(self, parts: List[str],
                             owner_lookup: Dict, class_lookup: Dict):
        """Helper to process a single FIELD or METHOD line."""
        keyword, owner_obf, desc, obf, inter, yarn = parts
        owner_data = owner_lookup.get(self._convert_path(owner_obf))
        if not owner_data:
            return

        base_info = {
            'Mojmap Obfuscated Class': owner_data['Mojmap Obfuscated Class'],
            'Yarn Deobfuscated Class': owner_data['Yarn Deobfuscated Class'],
            'Yarn Intermediary Class': owner_data['Yarn Intermediary Class'],
            'Mojmap Deobfuscated Class': '',
        }
        if keyword == 'FIELD':
            base_info.update({
                'Mojmap Obfuscated Field': obf,
                'Yarn Intermediary Field': inter,
                'Yarn Deobfuscated Field': yarn,
                'Mojmap Deobfuscated Field': '',
                'Yarn Type': self._parse_descriptor_type(desc, class_lookup),
                'Mojmap Type': ''
            })
            self.all_yarn_fields.append(base_info)
        elif keyword == 'METHOD':
            ret_type, sig = self._parse_yarn_signature(desc, yarn,
                                                       class_lookup)
            base_info.update({
                'Mojmap Obfuscated Method': obf,
                'Yarn Intermediary Method': inter,
                'Yarn Deobfuscated Method': yarn,
                'Mojmap Deobfuscated Method': '',
                'Obfuscated Descriptor': desc,
                'Yarn Return Type': ret_type,
                'Yarn Signature': sig,
                'Mojmap Return Type': '',
                'Mojmap Signature': ''
            })
            self.all_yarn_methods.append(base_info)

    def _parse_yarn_signature(self, desc: str, name: str,
                              lookup: Dict) -> Tuple[str, str]:
        """Parses a Yarn signature using a class lookup."""
        try:
            match = re.match(r'\((.*)\)(.*)', desc)
            if not match:
                return "unknown", "()"
            param_part, return_part = match.groups()
            return_type = self._parse_descriptor_type(return_part, lookup)
            param_types = DESCRIPTOR_RE.findall(param_part)
            param_list = [f"{self._parse_descriptor_type(pt, lookup)} arg{i}"
                          for i, pt in enumerate(param_types)]
            return return_type, f"{name}({', '.join(param_list)})"
        except Exception:
            return "error", "()"

    def merge_with_mojmap(self, mojmap_content: str):
        """Parses and merges Mojmap data, tracking used keys."""
        if not mojmap_content:
            print("No Mojmap content provided. Skipping merge.")
            return

        print("Parsing and merging Mojmap data...")
        self.mojmap_classes, self.mojmap_fields, self.mojmap_methods = \
            self._parse_mojmap_proguard_file(mojmap_content)

        # Merge classes and track usage
        for item in self.all_yarn_classes:
            key = item['Mojmap Obfuscated Class']
            mojmap_deobf = self.mojmap_classes.get(key, '')
            if mojmap_deobf:
                item['Mojmap Deobfuscated Class'] = mojmap_deobf
                self.used_mojmap_class_keys.add(key)

        obf_to_mojmap_owner = {
            c['Mojmap Obfuscated Class']: c['Mojmap Deobfuscated Class']
            for c in self.all_yarn_classes
        }

        # Merge fields and track usage
        for item in self.all_yarn_fields:
            item['Mojmap Deobfuscated Class'] = obf_to_mojmap_owner.get(
                item['Mojmap Obfuscated Class'], '')
            key = (item['Mojmap Obfuscated Class'],
                   item['Mojmap Obfuscated Field'])
            full_field = self.mojmap_fields.get(key, '')
            if full_field:
                self.used_mojmap_field_keys.add(key)
                parts = full_field.split()
                item['Mojmap Deobfuscated Field'] = parts[-1]
                item['Mojmap Type'] = self._convert_path(" ".join(parts[:-1]))

        # Merge methods and track usage
        for item in self.all_yarn_methods:
            item['Mojmap Deobfuscated Class'] = obf_to_mojmap_owner.get(
                item['Mojmap Obfuscated Class'], '')
            key = (item['Mojmap Obfuscated Class'],
                   item['Mojmap Obfuscated Method'],
                   item['Obfuscated Descriptor'])
            full_sig = self.mojmap_methods.get(key, '')
            if full_sig:
                self.used_mojmap_method_keys.add(key)
                item['Mojmap Deobfuscated Method'] = \
                    full_sig.split('(', 1)[0].split()[-1]
                ret_type, sig = self._parse_mojmap_signature(full_sig)
                item['Mojmap Return Type'] = ret_type
                item['Mojmap Signature'] = sig
        print("Merging complete.")

    def _parse_mojmap_proguard_file(self, content: str) -> Tuple:
        """Parses a Mojmap ProGuard file using a two-pass approach."""
        lines = content.splitlines()
        class_map, field_map, method_map = {}, {}, {}

        print("Mojmap Pass 1: Building class map...")
        for line in lines:
            stripped = line.strip()
            if not stripped or stripped.startswith('#'):
                continue
            processed = LINE_NUMBER_PREFIX_RE.sub('', stripped)
            if processed.endswith(':'):
                parts = processed.rstrip(':').split(' -> ')
                if len(parts) == 2:
                    mojmap_name = self._convert_path(parts[0])
                    obf_name = self._convert_path(parts[1])
                    class_map[obf_name] = mojmap_name

        mojmap_to_obf_map = {v: k for k, v in class_map.items()}

        print("Mojmap Pass 2: Building field and method maps...")
        current_obf_class = ""
        for line in lines:
            stripped = line.strip()
            if not stripped or stripped.startswith('#'):
                continue
            processed = LINE_NUMBER_PREFIX_RE.sub('', stripped)
            if processed.endswith(':'):
                current_obf_class = \
                    self._convert_path(processed.split(' -> ')[1].rstrip(':'))
            elif line.startswith(' ') and current_obf_class:
                parts = processed.split(' -> ')
                if len(parts) == 2:
                    deobf_sig, obf_name = parts[0].strip(), parts[1].strip()
                    if '(' in deobf_sig and ')' in deobf_sig:
                        desc = self._gen_obf_descriptor(deobf_sig,
                                                        mojmap_to_obf_map)
                        method_map[(current_obf_class, obf_name, desc)] = \
                            deobf_sig
                    else:
                        field_map[(current_obf_class, obf_name)] = deobf_sig
        return class_map, field_map, method_map

    def _gen_obf_descriptor(self, mojmap_sig: str,
                            mojmap_to_obf: Dict) -> str:
        """Generates a JNI-style descriptor from a deobfuscated signature."""
        try:
            primitive_map = {
                'byte': 'B', 'char': 'C', 'double': 'D', 'float': 'F',
                'int': 'I', 'long': 'J', 'short': 'S', 'boolean': 'Z',
                'void': 'V'
            }

            def type_to_desc(type_str: str) -> str:
                array_dims = type_str.count('[]')
                clean_type = type_str.replace('[]', '').strip()
                prefix = '[' * array_dims
                if clean_type in primitive_map:
                    return prefix + primitive_map[clean_type]
                obf_path = mojmap_to_obf.get(clean_type,
                                             clean_type.replace('.', '/'))
                return f"{prefix}L{obf_path.replace('.', '/')};"

            ret_type, method_params = self._parse_mojmap_signature(mojmap_sig)
            return_desc = type_to_desc(ret_type)
            params_str = method_params[method_params.find('(') + 1:-1]
            if not params_str:
                return f"(){return_desc}"

            param_types = [p.strip().split(' ')[0]
                           for p in params_str.split(',')]
            params_desc = "".join(type_to_desc(pt) for pt in param_types)
            return f"({params_desc}){return_desc}"
        except Exception:
            return "()?"

    # --- Data Segregation Methods ---
    def get_mapped_entries(self) -> Tuple[List, List, List]:
        """Returns lists of all fully matched classes, fields, and methods."""
        mapped_classes = [
            c for c in self.all_yarn_classes
            if c['Mojmap Deobfuscated Class']]
        mapped_fields = [
            f for f in self.all_yarn_fields
            if f['Mojmap Deobfuscated Field']]
        mapped_methods = [
            m for m in self.all_yarn_methods
            if m['Mojmap Deobfuscated Method']]
        return mapped_classes, mapped_fields, mapped_methods

    def get_unmatched_yarn_entries(self) -> Tuple[List, List, List]:
        """Returns Yarn entries that did not get a Mojmap match."""
        unmatched_classes = [
            c for c in self.all_yarn_classes
            if not c['Mojmap Deobfuscated Class']]
        unmatched_fields = [
            f for f in self.all_yarn_fields
            if not f['Mojmap Deobfuscated Field']]
        unmatched_methods = [
            m for m in self.all_yarn_methods
            if not m['Mojmap Deobfuscated Method']]
        return unmatched_classes, unmatched_fields, unmatched_methods

    def get_unmatched_mojmap_entries(self) -> Tuple[List, List, List]:
        """Returns Mojmap entries that were not used by any Yarn mapping."""
        unmatched_class_keys = \
            self.mojmap_classes.keys() - self.used_mojmap_class_keys
        unmatched_field_keys = \
            self.mojmap_fields.keys() - self.used_mojmap_field_keys
        unmatched_method_keys = \
            self.mojmap_methods.keys() - self.used_mojmap_method_keys

        unmatched_classes = [
            {
                'Mojmap Obfuscated Class': k,
                'Mojmap Deobfuscated Class': self.mojmap_classes[k]
            } for k in sorted(list(unmatched_class_keys))
        ]
        unmatched_fields = [
            {
                'Mojmap Obfuscated Class': k[0],
                'Mojmap Obfuscated Field': k[1],
                'Mojmap Deobfuscated Signature': self.mojmap_fields[k]
            } for k in sorted(list(unmatched_field_keys))
        ]
        unmatched_methods = [
            {
                'Mojmap Obfuscated Class': k[0],
                'Mojmap Obfuscated Method': k[1],
                'Obfuscated Descriptor': k[2],
                'Mojmap Deobfuscated Signature': self.mojmap_methods[k]
            } for k in sorted(list(unmatched_method_keys))
        ]
        return unmatched_classes, unmatched_fields, unmatched_methods

    # --- Data Filtering Methods ---
    def filter_client_packages(self):
        """Removes data from client-only packages."""
        print(f"\nFiltering out client packages: {CLIENT_PACKAGE_PREFIXES}")
        initial_count = len(self.all_yarn_classes)
        self.all_yarn_classes = [
            c for c in self.all_yarn_classes
            if not c['Yarn Deobfuscated Class'].startswith(
                CLIENT_PACKAGE_PREFIXES)
        ]
        kept_names = {c['Yarn Deobfuscated Class'] for c in self.all_yarn_classes}
        self.all_yarn_fields = [
            f for f in self.all_yarn_fields
            if f['Yarn Deobfuscated Class'] in kept_names]
        self.all_yarn_methods = [
            m for m in self.all_yarn_methods
            if m['Yarn Deobfuscated Class'] in kept_names]
        print(f"Removed {initial_count - len(self.all_yarn_classes)} "
              "client classes.")

    def sort_all_data(self):
        """Sorts all data lists for improved readability."""
        print("\nSorting data...")
        self.all_yarn_classes.sort(
            key=lambda item: item['Yarn Deobfuscated Class'])
        self.all_yarn_fields.sort(
            key=lambda item: (
                item['Yarn Deobfuscated Class'],
                item['Yarn Deobfuscated Field']
            )
        )
        self.all_yarn_methods.sort(
            key=lambda item: (
                item['Yarn Deobfuscated Class'],
                item['Yarn Deobfuscated Method']
            )
        )
        print("Sorting complete.")

    def apply_percentage_filter(self, percentage: int):
        """Filters the dataset to a top percentage of classes."""
        if not (1 <= percentage <= 100):
            print("\nProcessing all server-side data.")
            return

        print(f"\nProcessing top {percentage}% of server-side classes...")
        num_classes = int(len(self.all_yarn_classes) * (percentage / 100.0))
        self.all_yarn_classes = self.all_yarn_classes[:num_classes]
        kept_names = {c['Yarn Deobfuscated Class']
                      for c in self.all_yarn_classes}
        self.all_yarn_fields = [
            f for f in self.all_yarn_fields
            if f['Yarn Deobfuscated Class'] in kept_names]
        self.all_yarn_methods = [
            m for m in self.all_yarn_methods
            if m['Yarn Deobfuscated Class'] in kept_names]
        print(
            f"Reduced dataset to: {len(self.all_yarn_classes)} classes, "
            f"{len(self.all_yarn_fields)} fields, "
            f"{len(self.all_yarn_methods)} methods."
        )


class ExcelReportGenerator:
    """Handles the creation and formatting of the Excel spreadsheet."""

    def __init__(self, filename: str):
        """Initializes the workbook and color manager."""
        self.filename = filename
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.color_manager = ColorManager(COLOR_PALETTE)

    def generate_report(self, processor: MappingProcessor):
        """Creates the final Excel file from the processed data."""
        print(f"\nCreating Excel spreadsheet: {self.filename}")
        try:
            self._prepare_colors(processor.all_yarn_classes)

            # --- Mapped Entries ---
            mapped_c, mapped_f, mapped_m = processor.get_mapped_entries()
            self._write_mapped_class_sheet(mapped_c)
            self._write_mapped_field_sheet(mapped_f)
            self._write_mapped_method_sheet(mapped_m)

            # --- Unmatched Yarn Entries ---
            un_yarn_c, un_yarn_f, un_yarn_m = \
                processor.get_unmatched_yarn_entries()
            self._write_unmatched_yarn_class_sheet(un_yarn_c)
            self._write_unmatched_yarn_field_sheet(un_yarn_f)
            self._write_unmatched_yarn_method_sheet(un_yarn_m)

            # --- Unmatched Mojmap Entries ---
            un_mojmap_c, un_mojmap_f, un_mojmap_m = \
                processor.get_unmatched_mojmap_entries()
            self._write_unmatched_mojmap_class_sheet(un_mojmap_c)
            self._write_unmatched_mojmap_field_sheet(un_mojmap_f)
            self._write_unmatched_mojmap_method_sheet(un_mojmap_m)

            self._finalize_and_save()
        except PermissionError:
            print(f"\nError: Permission denied. Could not save "
                  f"'{self.filename}'.")
        except Exception as e:
            print(f"\nAn error occurred while creating the Excel file: {e}")

    def _prepare_colors(self, classes: List[Dict]):
        """Pre-populates the color manager for consistent coloring."""
        for cls_entry in classes:
            self.color_manager.get_fill_for_class(
                cls_entry.get('Yarn Deobfuscated Class') or
                cls_entry.get('Mojmap Deobfuscated Class')
            )

    # --- Writers for Mapped Sheets ---
    def _write_mapped_class_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Mojmap Deobfuscated Class", "Yarn Deobfuscated Class"
        ]
        self._write_sheet("Mapped Classes", headers, data,
                          "Yarn Deobfuscated Class")

    def _write_mapped_field_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Mojmap Deobfuscated Class", "Yarn Deobfuscated Class",
            "Mojmap Obfuscated Field", "Yarn Intermediary Field",
            "Mojmap Deobfuscated Field", "Yarn Deobfuscated Field",
            "Mojmap Type", "Yarn Type"
        ]
        self._write_sheet("Mapped Fields", headers, data,
                          "Yarn Deobfuscated Class")

    def _write_mapped_method_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Mojmap Deobfuscated Class", "Yarn Deobfuscated Class",
            "Mojmap Obfuscated Method", "Yarn Intermediary Method",
            "Mojmap Deobfuscated Method", "Yarn Deobfuscated Method",
            "Mojmap Return Type", "Yarn Return Type",
            "Mojmap Signature", "Yarn Signature"
        ]
        self._write_sheet("Mapped Methods", headers, data,
                          "Yarn Deobfuscated Class")

    # --- Writers for Unmatched Yarn Sheets ---
    def _write_unmatched_yarn_class_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Yarn Deobfuscated Class"
        ]
        self._write_sheet("Yarn Classes w_o Mojmap", headers, data,
                          "Yarn Deobfuscated Class")

    def _write_unmatched_yarn_field_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Yarn Deobfuscated Class", "Mojmap Obfuscated Field",
            "Yarn Intermediary Field", "Yarn Deobfuscated Field", "Yarn Type"
        ]
        self._write_sheet("Yarn Fields w_o Mojmap", headers, data,
                          "Yarn Deobfuscated Class")

    def _write_unmatched_yarn_method_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Yarn Intermediary Class",
            "Yarn Deobfuscated Class", "Mojmap Obfuscated Method",
            "Yarn Intermediary Method", "Yarn Deobfuscated Method",
            "Yarn Signature"
        ]
        self._write_sheet("Yarn Methods w_o Mojmap", headers, data,
                          "Yarn Deobfuscated Class")

    # --- Writers for Unmatched Mojmap Sheets ---
    def _write_unmatched_mojmap_class_sheet(self, data: List[Dict]):
        headers = ["Mojmap Obfuscated Class", "Mojmap Deobfuscated Class"]
        self._write_sheet("Mojmap Classes w_o Yarn", headers, data,
                          "Mojmap Deobfuscated Class")

    def _write_unmatched_mojmap_field_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Mojmap Obfuscated Field",
            "Mojmap Deobfuscated Signature"
        ]
        self._write_sheet("Mojmap Fields w_o Yarn", headers, data,
                          "Mojmap Obfuscated Class")

    def _write_unmatched_mojmap_method_sheet(self, data: List[Dict]):
        headers = [
            "Mojmap Obfuscated Class", "Mojmap Obfuscated Method",
            "Obfuscated Descriptor", "Mojmap Deobfuscated Signature"
        ]
        self._write_sheet("Mojmap Methods w_o Yarn", headers, data,
                          "Mojmap Obfuscated Class")

    # --- Generic Sheet Writing Helper ---
    def _write_sheet(self, name: str, headers: List[str], data: List[Dict],
                     row_fill_key: str):
        """Creates, formats, and populates a single worksheet."""
        print(f"Writing {len(data)} rows to '{name}' sheet...")
        sheet = self.workbook.create_sheet(name)
        sheet.append(headers)
        self._format_header_row(sheet.max_row, sheet)
        sheet.freeze_panes = 'A2'

        for row_data in data:
            fill = self.color_manager.get_fill_for_class(
                row_data.get(row_fill_key)
            )
            row_to_append = [row_data.get(key, '') for key in headers]
            sheet.append(row_to_append)
            if fill:
                for cell in sheet[sheet.max_row]:
                    cell.fill = fill

    def _format_header_row(self, row_index: int, sheet: Worksheet):
        """Applies standard formatting to a header row."""
        header_row = sheet[row_index]
        header_font = Font(bold=True, color="FFFFFF")
        header_border = Border(bottom=Side(style='thin', color='000000'))
        header_align = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        sheet.row_dimensions[row_index].height = 35

        for cell in header_row:
            cell.font = header_font
            cell.border = header_border
            cell.alignment = header_align
            color = HEADER_TYPE_COLORS["Generic"]
            if "Class" in cell.value:
                color = HEADER_TYPE_COLORS["Class"]
            elif "Field" in cell.value:
                color = HEADER_TYPE_COLORS["Field"]
            elif "Method" in cell.value:
                color = HEADER_TYPE_COLORS["Method"]
            cell.fill = PatternFill(start_color=color, end_color=color,
                                    fill_type="solid")

    def _finalize_and_save(self):
        """Auto-sizes columns and saves the workbook."""
        print("\nAuto-sizing columns for optimal spacing...")
        if "Mapped Classes" in self.workbook.sheetnames:
            self.workbook.active = self.workbook["Mapped Classes"]

        # The first sheet is a default one we don't use
        if len(self.workbook.sheetnames) > 1:
            del self.workbook[self.workbook.sheetnames[0]]

        for sheet in self.workbook.worksheets:
            self._auto_size_columns(sheet)
        print("Column sizing complete.")

        self.workbook.save(self.filename)
        print(f"\nSuccessfully created '{os.path.abspath(self.filename)}'")

    @staticmethod
    def _auto_size_columns(sheet: Worksheet):
        """Adjusts column widths based on content."""
        column_widths = defaultdict(int)
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                if cell.value:
                    max_line_len = max(len(line) for line
                                       in str(cell.value).split('\n'))
                    column_widths[i] = max(column_widths[i], max_line_len)

        for i, width in column_widths.items():
            adjusted_width = min(width + 2, 60)
            sheet.column_dimensions[get_column_letter(i + 1)].width = \
                adjusted_width


def get_file_content(prompt_msg: str, default_path: str,
                     optional: bool = False) -> Optional[str]:
    """Prompts user for a file path and returns its content."""
    prompt_suffix = "(or press Enter for default):"
    if optional:
        prompt_suffix = "(Optional, press Enter to skip):"

    print(f"{prompt_msg} {prompt_suffix}\n> {default_path}")
    file_path = input().strip() or default_path

    if not file_path and optional:
        return None
    try:
        print(f"Reading from: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        print(f"\nError: File not found at '{file_path}'")
        return None
    except Exception as e:
        print(f"\nError reading file '{file_path}': {e}")
        return None


def get_data_percentage() -> int:
    """Prompts the user for a percentage of data to process."""
    while True:
        p_input = input(
            "\nEnter data percentage to process (1-100), "
            "or -1 for all [default: -1]: "
        ).strip()
        if not p_input or p_input == '-1':
            return -1
        try:
            percentage = int(p_input)
            if 1 <= percentage <= 100:
                return percentage
            print("Invalid range. Please enter a number between 1-100.")
        except ValueError:
            print("Invalid input. Please enter a number.")


def main():
    """Main script execution flow."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    yarn_content = get_file_content(
        "Enter path to YARN .tiny mapping file", DEFAULT_YARN_TINY_PATH
    )
    if not yarn_content:
        sys.exit(1)

    processor = MappingProcessor()
    processor.parse_yarn_tiny_file(yarn_content)

    mojmap_content = get_file_content(
        "Enter path to MOJMAP server.txt mapping file", DEFAULT_MOJMAP_PATH,
        optional=True
    )
    if mojmap_content:
        processor.merge_with_mojmap(mojmap_content)

    processor.filter_client_packages()
    processor.sort_all_data()
    percentage = get_data_percentage()
    processor.apply_percentage_filter(percentage)

    version_part = re.search(r'(\d+\.\d+(\.\d+)?)',
                             os.path.basename(DEFAULT_YARN_TINY_PATH))
    base_name = (
        f"minecraft-mappings-by-category-{version_part.group(1)}.xlsx"
        if version_part else DEFAULT_OUTPUT_FILENAME
    )

    print(f"\nThe output file will be saved in: {script_dir}")
    filename_input = input(
        f"Enter output filename (or press Enter for '{base_name}'): "
    ).strip() or base_name
    if not filename_input.endswith('.xlsx'):
        filename_input += '.xlsx'
    final_path = os.path.join(script_dir, filename_input)

    reporter = ExcelReportGenerator(final_path)
    reporter.generate_report(processor)


if __name__ == "__main__":
    main()