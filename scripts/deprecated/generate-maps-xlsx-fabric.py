import os
import re
import itertools
from collections import defaultdict
from typing import List, Dict, Tuple, Any

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# --- Configuration ---
DEFAULT_INPUT_DIR = "."
# Make the default output filename more dynamic
DEFAULT_OUTPUT_FILENAME = "yarn_mappings_organized.xlsx"
PROGRESS_UPDATE_INTERVAL = 10000

# --- Regex & Mappings for Descriptors ---
DESCRIPTOR_RE = re.compile(r'(\[*(?:[BCDFIJSZ]|L[\w/$]+;))')
DESCRIPTOR_TYPE_MAP = {
    'B': 'byte', 'C': 'char', 'D': 'double', 'F': 'float',
    'I': 'int', 'J': 'long', 'S': 'short', 'Z': 'boolean',
    'V': 'void'
}

COLOR_PALETTE = [
    "FFCCBC", "E0F2F7", "E8F5E9", "FFF8E1", "F3E5F5", "FCE4EC",
    "CFD8DC", "D7CCC8", "DCEDC8", "B2EBF2", "FFE0B2", "D1C4E9",
    "C8E6C9", "FFCDD2", "BBDEFB", "FFFDE7", "F0F4C3", "ECCBCC"
]


class ColorManager:
    def __init__(self, palette: List[str]):
        self._palette_cycler = itertools.cycle(palette)
        self._class_fill_map: Dict[str, PatternFill] = {}

    def get_fill_for_class(self, class_name: str) -> PatternFill:
        if class_name not in self._class_fill_map:
            color_hex = next(self._palette_cycler)
            self._class_fill_map[class_name] = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )
        return self._class_fill_map[class_name]

def convert_path_to_dot_notation(path: str) -> str:
    return path.replace('/', '.')

def parse_descriptor_type(desc: str) -> str:
    array_suffix = ''
    while desc.startswith('['):
        array_suffix += '[]'
        desc = desc[1:]
    if desc.startswith('L') and desc.endswith(';'):
        type_name = convert_path_to_dot_notation(desc[1:-1])
    else:
        type_name = DESCRIPTOR_TYPE_MAP.get(desc, 'unknown')
    return type_name + array_suffix

def parse_signature_from_descriptor(descriptor: str, arg_names: List[str]) -> Tuple[str, str]:
    try:
        match = re.match(r'\((.*)\)(.*)', descriptor)
        if not match: return "unknown", "()"
        param_part, return_part = match.groups()
        return_type = parse_descriptor_type(return_part)
        param_types = DESCRIPTOR_RE.findall(param_part)
        param_list = []
        for i, p_type in enumerate(param_types):
            type_name = parse_descriptor_type(p_type)
            arg_name = arg_names[i] if i < len(arg_names) else f'arg{i}'
            param_list.append(f'{type_name} {arg_name}')
        return return_type, f"({', '.join(param_list)})"
    except Exception:
        return "error", "()"

def parse_member_line(parts: List[str]) -> Tuple[str, str, str]:
    if len(parts) == 2:
        return parts[0], parts[0], parts[1]
    elif len(parts) == 3:
        return parts[0], parts[1], parts[2]
    return "error", "error", ""

def parse_yarn_mappings(file_paths: List[str]) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    classes_data, fields_data, methods_data = [], [], []
    total_lines = 0
    file_count = len(file_paths)
    print(f"Parsing {file_count} mapping files...")

    for file_idx, file_path in enumerate(file_paths):
        # Optional: Add progress for files if you have many thousands
        # if (file_idx + 1) % 1000 == 0:
        #     print(f"  ...on file {file_idx + 1}/{file_count}")
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        class_stack = []
        last_item = None

        for line in lines:
            stripped_line = line.strip()
            if not stripped_line:
                continue

            # Indentation is still based on tabs, as that's part of the spec
            indentation = len(line) - len(line.lstrip('\t'))
            while len(class_stack) > indentation:
                class_stack.pop()
                # After popping, the last item context is the new top of the stack
                last_item = class_stack[-1] if class_stack else None

            # ************************************************************
            # THE FIX: Split by any whitespace, not just tabs.
            # ************************************************************
            parts = stripped_line.split()
            if not parts:
                continue

            keyword = parts[0]

            owner_intermediary = class_stack[-1]['Intermediary Name'] if class_stack else ""
            owner_named = class_stack[-1]['Named Name'] if class_stack else ""

            if keyword == 'CLASS':
                intermediary_name = convert_path_to_dot_notation(parts[1])
                named_name = convert_path_to_dot_notation(parts[2] if len(parts) > 2 else parts[1])
                class_info = {'Owner (Named)': owner_named, 'Intermediary Name': intermediary_name, 'Named Name': named_name, 'Comment': ''}
                classes_data.append(class_info)
                class_stack.append(class_info)
                last_item = class_info

            elif keyword == 'FIELD' and class_stack:
                intermediary, named, descriptor = parse_member_line(parts[1:])
                field_info = {'Owner (Intermediary)': owner_intermediary, 'Owner (Named)': owner_named, 'Intermediary Name': intermediary, 'Named Name': named, 'Type': parse_descriptor_type(descriptor), 'Comment': ''}
                fields_data.append(field_info)
                last_item = field_info

            elif keyword == 'METHOD' and class_stack:
                intermediary, named, descriptor = parse_member_line(parts[1:])
                method_info = {'Owner (Intermediary)': owner_intermediary, 'Owner (Named)': owner_named, 'Intermediary Name': intermediary, 'Named Name': named, 'Return Type': '', 'Signature': '', 'Raw Descriptor': descriptor, '_arg_names': [], 'Comment': ''}
                methods_data.append(method_info)
                last_item = method_info

            elif keyword == 'ARG' and last_item and '_arg_names' in last_item:
                arg_name = parts[2]
                last_item['_arg_names'].append(arg_name)

            elif keyword == 'COMMENT' and last_item:
                # Robustly get the full comment text after the keyword
                comment_parts = stripped_line.split(None, 1)
                comment_text = comment_parts[1] if len(comment_parts) > 1 else ""
                if last_item['Comment']:
                    last_item['Comment'] += '\n' + comment_text
                else:
                    last_item['Comment'] = comment_text

    print("Finalizing method signatures...")
    for method in methods_data:
        return_type, signature = parse_signature_from_descriptor(method['Raw Descriptor'], method['_arg_names'])
        method['Return Type'] = return_type
        method['Signature'] = f"{method['Named Name']}{signature}"
        del method['Raw Descriptor']
        del method['_arg_names']

    print("Parsing complete.")
    return classes_data, fields_data, methods_data


def setup_sheet(workbook: Workbook, sheet_name: str, headers: List[str]) -> Worksheet:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
    header_border = Border(bottom=Side(style='thin', color='000000'))
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
    sheet.freeze_panes = 'A2'
    return sheet

def write_data_to_sheet(sheet: Worksheet, data: List[Dict[str, Any]], color_manager: ColorManager, row_fill_key: str):
    headers = [cell.value for cell in sheet[1]]
    total_rows = len(data)
    print(f"Writing {total_rows} rows to '{sheet.title}' sheet...")
    max_widths = defaultdict(int)
    for col_idx, header_text in enumerate(headers):
        max_widths[col_idx] = len(str(header_text))
    for i, row_data in enumerate(data):
        color_key_value = row_data.get(row_fill_key) or row_data.get('Named Name', '')
        fill = color_manager.get_fill_for_class(color_key_value)
        row_to_append = [row_data.get(key, '') for key in headers]

        for col_idx, value in enumerate(row_to_append):
            current_width = len(str(value))
            if current_width > max_widths[col_idx]:
                max_widths[col_idx] = current_width

        sheet.append(row_to_append)
        if fill:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill
    auto_size_columns(sheet, max_widths)
    print("Writing complete.")

def auto_size_columns(sheet: Worksheet, max_widths: Dict[int, int]):
    for col_idx, max_width in max_widths.items():
        adjusted_width = min(max_width + 2, 80)
        col_letter = get_column_letter(col_idx + 1)
        sheet.column_dimensions[col_letter].width = adjusted_width

def create_excel_spreadsheet(classes: List[Dict], fields: List[Dict], methods: List[Dict], filename: str):
    print(f"\nCreating Excel spreadsheet: {filename}")
    try:
        workbook = openpyxl.Workbook()
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        color_manager = ColorManager(COLOR_PALETTE)
        for cls_entry in classes:
            color_manager.get_fill_for_class(cls_entry['Named Name'])

        class_headers = ["Owner (Named)", "Intermediary Name", "Named Name", "Comment"]
        classes_sheet = setup_sheet(workbook, "Classes", class_headers)
        write_data_to_sheet(classes_sheet, classes, color_manager, "Owner (Named)")

        field_headers = ["Owner (Named)", "Owner (Intermediary)", "Intermediary Name", "Named Name", "Type", "Comment"]
        fields_sheet = setup_sheet(workbook, "Fields", field_headers)
        write_data_to_sheet(fields_sheet, fields, color_manager, "Owner (Named)")

        method_headers = ["Owner (Named)", "Owner (Intermediary)", "Intermediary Name", "Named Name", "Return Type", "Signature", "Comment"]
        methods_sheet = setup_sheet(workbook, "Methods", method_headers)
        write_data_to_sheet(methods_sheet, methods, color_manager, "Owner (Named)")

        workbook.save(filename)
        print(f"\nSuccessfully created '{filename}'")
    except Exception as e:
        print(f"\nError creating Excel spreadsheet: {e}")

def main():
    input_dir = input(f"Enter the path to the mappings directory (or press Enter for '{DEFAULT_INPUT_DIR}'): ").strip()
    if not input_dir:
        input_dir = DEFAULT_INPUT_DIR

    if not os.path.isdir(input_dir):
        print(f"Error: Directory not found at '{input_dir}'")
        return

    # Extract version from path for a sensible default filename, if possible
    base_output_name = DEFAULT_OUTPUT_FILENAME
    path_parts = os.path.normpath(input_dir).split(os.sep)
    # Try to find a version number like '1.21.6' in the path
    version_part = next((part for part in reversed(path_parts) if re.match(r'^\d+\.\d+(\.\d+)?$', part)), None)
    if version_part:
        base_output_name = f"yarn-mappings-{version_part}.xlsx"

    output_filename = input(f"Enter the output Excel filename (or press Enter for '{base_output_name}'): ").strip()
    if not output_filename:
        output_filename = base_output_name
    if not output_filename.endswith('.xlsx'):
        output_filename += '.xlsx'

    mapping_files = []
    print(f"Searching for .mapping files in '{input_dir}'...")
    for root, _, files in os.walk(input_dir):
        for file in files:
            if file.endswith('.mapping'):
                mapping_files.append(os.path.join(root, file))

    if not mapping_files:
        print(f"Error: No '.mapping' files found in '{input_dir}' or its subdirectories.")
        return

    classes, fields, methods = parse_yarn_mappings(mapping_files)
    print(
        f"\nFinished parsing. Found:\n"
        f"  - {len(classes)} classes\n"
        f"  - {len(fields)} fields\n"
        f"  - {len(methods)} methods"
    )
    create_excel_spreadsheet(classes, fields, methods, output_filename)

if __name__ == "__main__":
    main()