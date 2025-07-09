import os
import re
import itertools
from collections import defaultdict
from typing import List, Dict, Tuple, Any

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import sys

MAPPINGS_URL = (
    "https://piston-data.mojang.com/v1/objects/"
    "94d453080a58875d3acc1a9a249809767c91ed40/server.txt"
)
INPUT_FILENAME = "server.txt"
OUTPUT_FILENAME = "minecraft_mappings_organized.xlsx"
PROGRESS_UPDATE_INTERVAL = 5000  # Number of lines/rows after which to print progress

# --- REGEX Definitions ---
# Regex to remove optional line numbers like "1:2:" or "1:" from the start of a line
LINE_NUMBER_PREFIX_RE = re.compile(r"^(?:\d+:\d*:|\d+:)")

# This regex is specifically for method names like <init> or <clinit>
SPECIAL_METHOD_NAME_RE = re.compile(r"^(?P<type>[\w./$<>\[\]]+)?\s*(?P<name><init>|<clinit>)(?P<params>\(.*\))?$")

COLOR_PALETTE = [
    "FFCCBC", "E0F2F7", "E8F5E9", "FFF8E1", "F3E5F5", "FCE4EC",
    "CFD8DC", "D7CCC8", "DCEDC8", "B2EBF2", "FFE0B2", "D1C4E9",
    "C8E6C9", "FFCDD2", "BBDEFB", "FFFDE7", "F0F4C3", "ECCBCC"
]


class ColorManager:
    def __init__(self, palette: List[str]):
        self._palette_cycler = itertools.cycle(palette)
        self._class_fill_map: Dict[str, PatternFill] = {}

    def get_fill_for_class(self, class_name: str) -> PatternFill:
        if class_name not in self._class_fill_map:
            color_hex = next(self._palette_cycler)
            self._class_fill_map[class_name] = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )
        return self._class_fill_map[class_name]

def convert_path_to_dot_notation(path: str) -> str:
    """Converts a Java-style path with slashes to dot notation."""
    return path.replace('/', '.')


def parse_mappings(raw_text: str) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    if not raw_text:
        return [], [], []

    classes_data, fields_data, methods_data = [], [], []
    current_class_deobf, current_class_obf = "", ""

    lines = raw_text.splitlines()
    total_lines = len(lines)
    print("Parsing mapping lines...")

    skipped_lines_count = 0

    for i, line in enumerate(lines):
        if (i + 1) % PROGRESS_UPDATE_INTERVAL == 0:
            print(f"  ...processed {i + 1}/{total_lines} lines")

        stripped_line = line.strip()
        if not stripped_line or stripped_line.startswith('#'):
            continue

        # --- Member vs. Class Detection based on indentation ---
        # Members are indented, classes are not.

        # ************************************************************
        # FIX: Check for indentation on the ORIGINAL line, not the stripped one.
        # ************************************************************
        if line.startswith(' '):  # This is a member line
            # It must belong to a class parsed previously
            if not current_class_deobf:
                skipped_lines_count += 1
                continue

            # Format:    deobfuscated_signature -> obfuscated_name
            processed_line = LINE_NUMBER_PREFIX_RE.sub('', stripped_line)
            parts = processed_line.split(' -> ', 1)

            if len(parts) == 2:
                deobf_signature = parts[0].strip()
                obf_member_name = parts[1].strip()

                # --- Member Type Differentiation (Method vs Field) ---
                if '(' in deobf_signature and ')' in deobf_signature:
                    # It's a method
                    method_name = ""
                    return_type = ""

                    special_match = SPECIAL_METHOD_NAME_RE.match(deobf_signature)
                    if special_match:
                        method_name = special_match.group('name')
                        return_type_raw = special_match.group('type')
                        return_type = convert_path_to_dot_notation(return_type_raw) if return_type_raw else ''
                    else:
                        parts_before_params = deobf_signature.split('(', 1)[0]
                        last_space_idx = parts_before_params.rfind(' ')
                        if last_space_idx != -1:
                            return_type_raw = parts_before_params[:last_space_idx].strip()
                            method_name = parts_before_params[last_space_idx+1:].strip()
                            return_type = convert_path_to_dot_notation(return_type_raw)
                        else:
                            method_name = parts_before_params.strip()
                            return_type = "" # No return type found

                    methods_data.append({
                        'Owner (Obfuscated)': current_class_obf,
                        'Owner (Named)': current_class_deobf,
                        'Method (Obfuscated)': obf_member_name,
                        'Method (Named)': method_name,
                        'Return Type': return_type,
                        'Signature': convert_path_to_dot_notation(deobf_signature)
                    })
                else:
                    # It's a field
                    last_space_idx = deobf_signature.rfind(' ')
                    if last_space_idx != -1:
                        field_type = convert_path_to_dot_notation(deobf_signature[:last_space_idx].strip())
                        field_name = deobf_signature[last_space_idx+1:].strip()
                    else:
                        field_type = "" # No type found
                        field_name = deobf_signature

                    fields_data.append({
                        'Owner (Obfuscated)': current_class_obf,
                        'Owner (Named)': current_class_deobf,
                        'Field (Obfuscated)': obf_member_name,
                        'Field (Named)': field_name,
                        'Type': field_type
                    })
            else:
                skipped_lines_count += 1

        else:  # This is a class line (not indented)
            # Format: obfuscated_path -> deobfuscated_path:
            processed_line = LINE_NUMBER_PREFIX_RE.sub('', stripped_line)
            if processed_line.endswith(':') and ' -> ' in processed_line:
                parts = processed_line.rstrip(':').split(' -> ')
                if len(parts) == 2:
                    current_class_obf = convert_path_to_dot_notation(parts[0])
                    current_class_deobf = convert_path_to_dot_notation(parts[1])

                    classes_data.append({
                        'Obfuscated Name': current_class_obf,
                        'Named Name': current_class_deobf
                    })
                else:
                    skipped_lines_count += 1
            else:
                skipped_lines_count += 1

    print("Parsing complete.")
    if skipped_lines_count > 0:
        print(f"Warning: Skipped {skipped_lines_count} lines due to unmatched format or missing class context.")
    return classes_data, fields_data, methods_data


def setup_sheet(workbook: Workbook, sheet_name: str, headers: List[str]) -> Worksheet:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
    header_border = Border(bottom=Side(style='thin', color='000000'))

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border

    sheet.freeze_panes = 'A2'
    return sheet


def write_data_to_sheet(
    sheet: Worksheet,
    data: List[Dict[str, Any]],
    color_manager: ColorManager,
    row_fill_key: str
):
    headers = [cell.value for cell in sheet[1]]
    total_rows = len(data)
    print(f"Writing {total_rows} rows to '{sheet.title}' sheet...")

    max_widths = defaultdict(int)
    for col_idx, header_text in enumerate(headers):
        max_widths[col_idx] = len(str(header_text))

    for i, row_data in enumerate(data):
        if (i + 1) % PROGRESS_UPDATE_INTERVAL == 0:
            print(f"  ...wrote {i + 1}/{total_rows} rows")

        fill = color_manager.get_fill_for_class(row_data.get(row_fill_key, ""))
        row_to_append = []
        for col_idx, key in enumerate(headers):
            value = row_data.get(key, '')
            row_to_append.append(value)
            current_width = len(str(value))
            if current_width > max_widths[col_idx]:
                max_widths[col_idx] = current_width

        sheet.append(row_to_append)
        if fill:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill

    auto_size_columns(sheet, max_widths)
    print("Writing complete.")


def auto_size_columns(sheet: Worksheet, max_widths: Dict[int, int]):
    for col_idx, max_width in max_widths.items():
        adjusted_width = min(max_width + 2, 80)
        col_letter = get_column_letter(col_idx + 1)
        sheet.column_dimensions[col_letter].width = adjusted_width


def create_excel_spreadsheet(
    classes: List[Dict], fields: List[Dict], methods: List[Dict], filename: str
):
    print(f"\nCreating Excel spreadsheet: {filename}")
    try:
        workbook = openpyxl.Workbook()
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        color_manager = ColorManager(COLOR_PALETTE)

        for cls_entry in classes:
            color_manager.get_fill_for_class(cls_entry['Named Name'])

        class_headers = ["Obfuscated Name", "Named Name"]
        classes_sheet = setup_sheet(workbook, "Classes", class_headers)
        write_data_to_sheet(classes_sheet, classes, color_manager, "Named Name")

        field_headers = [
            "Owner (Obfuscated)", "Owner (Named)", "Field (Obfuscated)",
            "Field (Named)", "Type"
        ]
        fields_sheet = setup_sheet(workbook, "Fields", field_headers)
        write_data_to_sheet(fields_sheet, fields, color_manager, "Owner (Named)")

        method_headers = [
            "Owner (Obfuscated)", "Owner (Named)", "Method (Obfuscated)",
            "Method (Named)", "Return Type", "Signature"
        ]
        methods_sheet = setup_sheet(workbook, "Methods", method_headers)
        write_data_to_sheet(methods_sheet, methods, color_manager, "Owner (Named)")

        workbook.save(filename)
        print(f"\nSuccessfully created '{filename}'")
    except Exception as e:
        print(f"\nError creating Excel spreadsheet: {e}")


def main():
    if not os.path.exists(INPUT_FILENAME):
        print(f"Input file '{INPUT_FILENAME}' not found. Downloading from Mojang...")
        try:
            r = requests.get(MAPPINGS_URL, timeout=60)
            r.raise_for_status()
            with open(INPUT_FILENAME, 'w', encoding='utf-8') as f:
                f.write(r.text)
            print("Download successful.")
        except requests.exceptions.RequestException as e:
            sys.exit(f"Error: Download failed: {e}.")
    else:
        print(f"Using existing input file: '{INPUT_FILENAME}'")

    with open(INPUT_FILENAME, "r", encoding="utf-8") as f:
        raw_mappings = f.read()

    if raw_mappings:
        classes, fields, methods = parse_mappings(raw_mappings)
        print(
            f"\nFinished parsing. Found:\n"
            f"  - {len(classes)} classes\n"
            f"  - {len(fields)} fields\n"
            f"  - {len(methods)} methods"
        )
        create_excel_spreadsheet(classes, fields, methods, OUTPUT_FILENAME)
    else:
        print("Could not proceed without mappings data.")


if __name__ == "__main__":
    main()