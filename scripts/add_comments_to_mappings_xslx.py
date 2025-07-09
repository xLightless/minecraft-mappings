import os
import re
import sys
from collections import defaultdict
from typing import Dict, List

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# --- Configuration ---
DEFAULT_FULL_YARN_DIR = r"C:\Users\xLightless\source\repos\xLightless\WorldManager\net.minecraft.maps\1.21.6\yarn\mappings"
DEFAULT_EXCEL_FILE = r"C:\Users\xLightless\source\repos\xLightless\WorldManager\net.minecraft.maps\scripts\vanilla-and-yarn\minecraft-mappings-mojmap-fabric-1.21.6.xlsx"
COMMENT_HEADER_COLOR = "9F94C4"  # A muted purple for the comment column header
COMMENT_CELL_COLOR = "EAE6F8"    # A very light purple for the comment data cells

def parse_full_yarn_mappings(yarn_dir: str) -> Dict[str, str]:
    """
    Parses the full Yarn directory to get comments for classes, fields,
    and methods, keyed by their intermediary names.
    """
    print(f"Parsing full Yarn mappings directory for comments...")
    details_map = defaultdict(str)
    mapping_files = [
        os.path.join(root, file)
        for root, _, files in os.walk(yarn_dir)
        for file in files if file.endswith('.mapping')
    ]
    if not mapping_files:
        print(f"Warning: No '.mapping' files found in '{yarn_dir}'.")
        return {}
    print(f"Found {len(mapping_files)} .mapping files to process.")

    for file_path in mapping_files:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        class_stack, last_item_key = [], None
        for line in lines:
            stripped = line.strip()
            if not stripped: continue
            indentation = len(line) - len(line.lstrip('\t'))
            while len(class_stack) > indentation: class_stack.pop()
            parts = stripped.split()
            keyword = parts[0]
            owner_intermediary_key = ".".join(class_stack)
            if keyword == 'CLASS':
                intermediary_name_path = parts[1].replace('/', '.')
                class_stack.append(intermediary_name_path)
                last_item_key = ".".join(class_stack)
            elif keyword == 'FIELD' and owner_intermediary_key:
                last_item_key = f"{owner_intermediary_key}.{parts[1]}"
            elif keyword == 'METHOD' and owner_intermediary_key:
                last_item_key = f"{owner_intermediary_key}.{parts[1]}"
            elif keyword == 'COMMENT' and last_item_key:
                comment_text = stripped.split(None, 1)[1] if len(parts) > 1 else ""
                existing = details_map[last_item_key]
                details_map[last_item_key] = f"{existing}\n{comment_text}" if existing else comment_text
    for key in details_map: details_map[key] = details_map[key].strip()
    print(f"Finished parsing comments from {len(details_map)} entries.")
    return details_map


def get_column_indices(sheet: Worksheet, headers: List[str]) -> Dict[str, int]:
    """Finds the 1-based column indices for a list of header names."""
    indices = {}
    header_row = [cell.value for cell in sheet[1]]
    for header in headers:
        try:
            indices[header] = header_row.index(header) + 1
        except ValueError:
            print(f"Warning: Header '{header}' not found in sheet '{sheet.title}'.")
            indices[header] = -1
    return indices


def add_comment_column(sheet: Worksheet):
    """Adds and formats a new 'Comment' column to a sheet."""
    comment_col_idx = sheet.max_column + 1
    header_cell = sheet.cell(row=1, column=comment_col_idx, value="Comment")
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(
        start_color=COMMENT_HEADER_COLOR, end_color=COMMENT_HEADER_COLOR,
        fill_type="solid"
    )
    header_cell.border = Border(bottom=Side(style='thin', color='000000'))
    header_cell.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True
    )
    return comment_col_idx


def apply_comments_and_formatting(workbook: Workbook, comment_map: Dict[str, str]):
    """Applies comments and formats cells in the workbook sheets."""
    print("\nApplying comments and formatting to Excel file...")
    comment_cell_fill = PatternFill(
        start_color=COMMENT_CELL_COLOR, end_color=COMMENT_CELL_COLOR,
        fill_type="solid"
    )
    # --- CHANGE: Define the alignment for data cells ---
    data_alignment = Alignment(wrap_text=True, vertical='top')

    # --- Enrich Classes Sheet ---
    print("  Updating 'Classes' sheet...")
    classes_sheet = workbook['Classes']
    class_cols = get_column_indices(classes_sheet, ["Yarn Intermediary Class"])
    comment_col_idx = add_comment_column(classes_sheet)
    for row in classes_sheet.iter_rows(min_row=2):
        for cell in row: cell.alignment = data_alignment # Apply wrap text
        row[comment_col_idx - 1].fill = comment_cell_fill
        intermediary_class = row[class_cols["Yarn Intermediary Class"] - 1].value
        if intermediary_class and intermediary_class in comment_map:
            row[comment_col_idx - 1].value = comment_map[intermediary_class]

    # --- Enrich Fields Sheet ---
    print("  Updating 'Fields' sheet...")
    fields_sheet = workbook['Fields']
    field_cols = get_column_indices(
        fields_sheet, ["Yarn Intermediary Class", "Yarn Intermediary Field"]
    )
    comment_col_idx = add_comment_column(fields_sheet)
    for row in fields_sheet.iter_rows(min_row=2):
        for cell in row: cell.alignment = data_alignment # Apply wrap text
        row[comment_col_idx - 1].fill = comment_cell_fill
        owner_inter = row[field_cols["Yarn Intermediary Class"] - 1].value
        field_inter = row[field_cols["Yarn Intermediary Field"] - 1].value
        if owner_inter and field_inter:
            key = f"{owner_inter}.{field_inter}"
            if key in comment_map:
                row[comment_col_idx - 1].value = comment_map[key]

    # --- Enrich Methods Sheet ---
    print("  Updating 'Methods' sheet...")
    methods_sheet = workbook['Methods']
    method_cols = get_column_indices(
        methods_sheet, ["Yarn Intermediary Class", "Yarn Intermediary Method"]
    )
    comment_col_idx = add_comment_column(methods_sheet)
    for row in methods_sheet.iter_rows(min_row=2):
        for cell in row: cell.alignment = data_alignment # Apply wrap text
        row[comment_col_idx - 1].fill = comment_cell_fill
        owner_inter = row[method_cols["Yarn Intermediary Class"] - 1].value
        method_inter = row[method_cols["Yarn Intermediary Method"] - 1].value
        if owner_inter and method_inter:
            key = f"{owner_inter}.{method_inter}"
            if key in comment_map:
                row[comment_col_idx - 1].value = comment_map[key]

    print("Enrichment complete.")


def auto_size_all_columns(sheet: Worksheet):
    """Auto-sizes all columns in a worksheet based on content."""
    column_widths = defaultdict(int)
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value:
                cell_lines = str(cell.value).split('\n')
                # Check if any line in the cell is longer than the current max
                # This respects manual newlines in comments
                max_line_length = max(len(line) for line in cell_lines)
                column_widths[i] = max(column_widths[i], max_line_length)

    for i, column_width in column_widths.items():
        # Set a generous max width, but allow smaller if content fits
        adjusted_width = min(column_width + 2, 50)
        sheet.column_dimensions[get_column_letter(i + 1)].width = \
            adjusted_width


def main():
    """Main execution flow for the enrichment script."""
    excel_path = input(
        "Enter the path to the Excel file to enrich\n"
        f"(or press Enter for '{DEFAULT_EXCEL_FILE}'): "
    ).strip() or DEFAULT_EXCEL_FILE

    if not os.path.exists(excel_path):
        print(f"Error: File not found at '{excel_path}'")
        sys.exit(1)

    yarn_dir = input(
        "Enter path to the FULL Yarn mappings directory\n"
        f"(or press Enter for '{DEFAULT_FULL_YARN_DIR}'): "
    ).strip() or DEFAULT_FULL_YARN_DIR

    if not os.path.isdir(yarn_dir):
        print(f"Error: Directory not found at '{yarn_dir}'")
        sys.exit(1)

    comment_map = parse_full_yarn_mappings(yarn_dir)
    if not comment_map:
        print("Could not parse any comments. Exiting.")
        return

    try:
        print(f"\nLoading workbook: {excel_path}")
        workbook = openpyxl.load_workbook(excel_path)

        apply_comments_and_formatting(workbook, comment_map)

        # --- CHANGE: Auto-size columns AFTER all data and formatting is applied ---
        print("\nAuto-sizing columns for optimal spacing...")
        for sheet in workbook.worksheets:
            auto_size_all_columns(sheet)
        print("Column sizing complete.")

        path_parts = os.path.splitext(excel_path)
        new_filename = f"{path_parts[0]}_enriched{path_parts[1]}"

        workbook.save(new_filename)
        print(f"\nSuccessfully created enriched file: '{os.path.abspath(new_filename)}'")

    except Exception as e:
        print(f"\nAn error occurred: {e}")


if __name__ == "__main__":
    main()